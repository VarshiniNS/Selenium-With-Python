import openpyxl


class HomePagedata:
    # Intially created a dictonary in the test date passed it
    # test_data_for_form_submission=[{"FirstName": "Apple", "Email": "fruits@gmail.com ", "Password": "Passs", "Gender": "Female"}
    #         , {"FirstName": "Carrot", "Email": "Vegetable@gmail.com ", "Password": "Veggie", "Gender": "Female"},
    #             {"FirstName": "rose", "Email": "flower@gmail.com ", "Password": "flowery", "Gender": "Male"}]

    #decalred as static so that no need of creating object to use in test file
    @staticmethod
    def getTestData(test_Case_Name):
        workbook = openpyxl.load_workbook(
            "C://Users//Varshini//PycharmProjects//pythonFrameWorkProject//testData//TestData.xlsx")
        sheet = workbook.active
        dict = {}

        for i in range(2, sheet.max_row + 1):
            if sheet.cell(row=i,column=1).value == test_Case_Name :
                for j in range(2, sheet.max_column + 1):
                    dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [dict]